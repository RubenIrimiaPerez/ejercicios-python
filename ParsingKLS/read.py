from openpyxl import load_workbook

wb = load_workbook('documento.xlsx')

ws = wb.active

for row in ws.iter_rows():
    for cell in row:
        print(cell.value)